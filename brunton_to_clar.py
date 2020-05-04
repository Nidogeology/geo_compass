import openpyxl

def get_and_transform_to_clar():
    strike_b = col1.value
    dip_b = col2.value
    global strike_c, ndip_c

    nstrike_b = ''
    ndip_b = ''
    b1 = strike_b[0]
    b1 = b1.upper()
    b2 = strike_b[-1]
    b2 = b2.upper()
    b3 = dip_b[-2:]
    b3 = b3.upper()

    for i in strike_b:
        if i.isnumeric():
            nstrike_b += i
    nstrike_c = int(nstrike_b)

    for i in dip_b:
        if i.isnumeric():
            ndip_b += i
    ndip_c = int(ndip_b)

    if b2 == 'W' and b3 == 'SW':
        strike_c = 270 - int(nstrike_b)
    elif b2 == 'W' and b3 == 'NE':
        strike_c = 90 - int(nstrike_b)
    elif b2 == 'E' and b3 == 'SE':
        strike_c = int(nstrike_b) + 90
    elif b2 == 'E' and b3 == 'NW':
        strike_c = 360 - int(nstrike_b)
    ndip_c = ndip_b

def write_clar():
    global index

    plan2.cell(index, 1).value = strike_c
    plan2.cell(index, 2).value = ndip_c
    index += 1


wb_brunton = openpyxl.load_workbook('brunton_data.xlsx')
plan1 = wb_brunton[wb_brunton.sheetnames[0]]
plan1.title = 'Brunton'
plan2 = wb_brunton.create_sheet('Clar')
index = 1

for col1, col2 in plan1:
    get_and_transform_to_clar()
    write_clar()

wb_brunton.save('brunton_to_clar_output.xlsx')


