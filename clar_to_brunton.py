import openpyxl


def get_and_transform_to_brunton():
    strike_c = col1.value
    dip_c = col2.value
    global strike_b, dip_b, b1, b2, b3

    dip_b = dip_c

    if strike_c == 0 or strike_c == 360:
        strike_b = 'E-W'
        b1 = ''
        b2 = ''
        b3 = 'N'
    if strike_c == 180:
        strike_b = 'E-W'
        b1 = ''
        b2 = ''
        b3 = 'S'
    elif strike_c == 90:
        strike_b = 'N-S'
        b1 = ''
        b2 = ''
        b3 = 'E'
    elif strike_c == 270:
        strike_b = 'N-S'
        b1 = ''
        b2 = ''
        b3 = 'W'
    elif 0 < strike_c < 90:
        strike_b = (90 - strike_c)
        b1 = 'N'
        b2 = 'W'
        b3 = 'NE'
    elif 90 < strike_c < 180:
        strike_b = strike_c - 90
        b1 = 'N'
        b2 = 'E'
        b3 = 'SE'
    elif 180 < strike_c < 270:
        strike_b = 360 - (strike_c + 90)
        b1 = 'N'
        b2 = 'W'
        b3 = 'SW'
    elif 270 < strike_c < 360:
        strike_b = (strike_c + 90) - 360
        b1 = 'N'
        b2 = 'E'
        b3 = 'NW'

    if dip_c == 0:
        strike_b = 'horizontal'
        dip_b = ''
        b3 = ''
    if dip_c == 90:
        dip_b = ''
        b3 = 'vertical'

def write_brunton():
    global index
    plan2.cell(index, 1).value = b1 + str(strike_b) + b2
    plan2.cell(index, 2).value = str(dip_b) + b3
    index += 1


wb_clar = openpyxl.load_workbook('clar_data.xlsx')

plan1 = wb_clar[wb_clar.sheetnames[0]]
plan1.title = 'Clar'
plan2 = wb_clar.create_sheet('Brunton')
index = 1

for col1, col2 in plan1:
    get_and_transform_to_brunton()
    write_brunton()


wb_clar.save('clar_to_brunton_output.xlsx')